# imports
from openpyxl import Workbook, load_workbook
import os

# function
def compile_workbooks(workbooks_path, final_filename):
    
    # if the path isnt a string
    if not isinstance(workbooks_path, str):
        raise TypeError("Argument workbooks_path must be of type str.")

    # if the file isnt a string
    if not isinstance(final_filename, str):
        raise TypeError("Argument final_filename must be of type str.")

    # if the path isnt a directory
    if not os.path.exists(workbooks_path):
        raise NotADirectoryError("Argument workbook_path is not a directory.")

    # if the file isnt an excel workbook
    if not final_filename.endswith(".xlsx"):
        raise ValueError('final_filename must end with the string ".xlsx"')
    
    # if the workbook name already exists in the path
    if final_filename in os.listdir(workbooks_path):
        raise ValueError(f'There is already a file named {final_filename} in {workbooks_path}. '
                         f'Remove this file first or change the final_filename parameter value.')
    
    # list that will save all the workbooks name that we will use
    wbs = []

    # for file in path, if it isnt a temporary file and is an excel workbook we will read and append in list
    for file in os.listdir(workbooks_path):
        if not file.startswith("~$") and file.endswith(".xlsx"):
            wb = load_workbook(os.path.join(workbooks_path, file))
            wbs.append(wb)
    
    # open the first workbook
    final_wb = Workbook()
    final_ws = final_wb.worksheets[0]
    wb1 = wbs[0]
    ws1 = wb1.worksheets[0] 

    # creating the header of the workbook
    for j in range(1, ws1.max_column+1):
        final_ws.cell(row=1, column=j).value = ws1.cell(row=1, column=j).value

    current_row = 2

    # for workbook in list
    for wb in wbs:
        # saving the max row and max colum for each one and compiling that
        for ws in wb.worksheets:
            mr = ws.max_row 
            mc = ws.max_column 

            for i in range (2, mr + 1): 
                for j in range (1, mc + 1): 
                    current_cell = ws.cell(row = i, column = j) 
                    final_ws.cell(row = current_row, column = j).value = current_cell.value

                current_row += 1

    # salving the final workbook
    final_wb.save(os.path.join(workbooks_path, final_filename))
